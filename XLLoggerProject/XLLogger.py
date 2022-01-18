#Week 1 weekend project - xl logger by Diego Suarez

from openpyxl import load_workbook
import os, logging

logging.basicConfig(filename = "XLLoggerProject/xllog.log", filemode = "w", format ="%(asctime)s - %(levelname)s %(message)s", datefmt = '%Y-%m-%d %H:%M:%S', encoding = 'utf-8', level = logging.DEBUG)
xlFiles = [file for file in os.listdir('XLLoggerProject/') if ".xlsx" in file]
months = {
    'january': '01', 'february': '02', 'march': '03', 'april': '04', 'may': '05', 'june': '06', 'july': '07', 'august': '08', 'september': '09', 'october': '10', 'november': '11', 'december': '12', 
    'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04', 'may': '05', 'jun': '06', 'jul': '07', 'aug': '08', 'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12'
    }

#display xl workbook files located within XLLoggerProject/ folder and allow user to select which file to parse, which is then returned
def getWorkbookName():
    if not xlFiles:
        msg = "No xlsx files found"
        logging.warning(msg)
        print(msg)
        exit()

    print("Please select the index of the file you would like to parse: ")
    for i, file in enumerate(xlFiles):
        print(i + 1, file)
    
    while True:
        try:
            choice = int(input())
        except ValueError:
            print("Invalid input.")
            continue

        if choice <= len(xlFiles) and choice != 0:
            print("Selected:", choice, xlFiles[choice-1])
            return xlFiles[choice-1][:-5]
        else:
            print("Invalid choice.")
        

"""def getDigit():
    n = input()

    if n.isdigit():
        return int(n)
    else:
        print("Invalid choice.")
        return getDigit()"""

#executes main program
def workbook_loader():
    workbook_name = getWorkbookName()
    workbook_path = "XLLoggerProject/" + workbook_name + ".xlsx"
    
    #verify file exists, otherwise exit program
    if not os.path.exists(workbook_path):
        msg = "File does not exist."
        logging.warning(msg)
        print(msg)
        exit()

    workbook = load_workbook(workbook_path)
    worksheet = workbook["Summary Rolling MoM"]

    #get the year and month of file by splitting and slicing up to the last 2 items in the list
    workbook_date = workbook_name.split("_")[-1:-3:-1]
    print(workbook_date)

    #format the year and month from previous list result into a format that would be included within the datetime value present in the workbook
    year_month = f'{workbook_date[0]}-{months[workbook_date[1].lower()]}'
    print(year_month)

    #find the row the matches the date of the file by iterating through the first column
    row = 0
    for cell in worksheet["A"]:
        row += 1
        if year_month in str(cell.value):
            print("Found match", year_month, cell.value, "row", row)
            break
    
    header_row = list(filter(lambda x : x != None, list(worksheet.iter_rows(max_row = 1, values_only = True))[0]))
    current_row = list(filter(lambda x : x != None, list(worksheet.iter_rows(min_row = row, max_row = row, values_only = True))[0][1:]))

    #print(list(worksheet.iter_rows(max_row = 1, values_only = True))[0])
    print(header_row)
    print(current_row)

    #log results from sheet 1
    log_worksheet1(workbook_date, header_row, current_row)

    #switch to second sheet
    worksheet = workbook["VOC Rolling MoM"]

    #find the column that matches the date of the file by iterating through the first row
    col = 0
    for cell in worksheet["1"]:
        col += 1
        if year_month in str(cell.value) or workbook_date[1].capitalize() in str(cell.value):
            print("Found match", year_month, cell.value, "col", col)
            break

    label_col = list(filter(lambda x : x != None, list(worksheet.iter_cols(max_col = 1, values_only = True))[0]))[:5]
    current_col = list(filter(lambda x : x != None and int(x) != 0, list(worksheet.iter_cols(min_col = col, max_col = col, values_only = True))[0][3:]))
    
    print(label_col)
    print(current_col, len(current_col))
    print(create_scores_list(current_col))
    
    log_workSheet2(label_col, create_scores_list(current_col))

#return list of tuples for each score with its evaluation of good or bad
def create_scores_list(scores):
    evaluations = []
    if scores[0] > 200:
        evaluations.append((scores[0], "Good"))
    else:
        evaluations.append((scores[0], "Bad"))
    if scores[1] > 100:
        evaluations.append((scores[1], "Good"))
    else:
        evaluations.append((scores[1], "Bad"))
    if scores[2] > 100:
        evaluations.append((scores[2], "Good"))
    else:
        evaluations.append((scores[2], "Bad"))
    return evaluations

def log_worksheet1(date, header_cells, values):
    logging.info(f"Month of {date[1].capitalize()}, {date[0]}" )
    logging.info(f"{header_cells[0]}: {values[0]}")
    for x in range(1, len(values)):
        logging.info(f"{header_cells[x]}: {values[x]*100}%")
    
def log_workSheet2(header, scores):
    logging.info(header[0])
    for x in range(0, len(scores)):
        logging.info(f"{header[x+2]}: {scores[x][0]} : {scores[x][1]}")
        


if __name__ == "__main__":
    workbook_loader()